import openpyxl
from django.http import HttpResponse
from datetime import datetime
from .models import Madre, Parto, RecienNacido, VacunaBCG



def remove_tz(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    return value



def auto_adjust(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2


def exportar_formularios_excel(madres_qs):
    wb = openpyxl.Workbook()

    # =========================
    # HOJA MADRE
    # =========================
    ws_madre = wb.active
    ws_madre.title = "Madre"

    headers_madre = [
        "id",
        "nombre_completo",
        "rut",
        "fecha_nacimiento",
        "edad",
        "comuna",
        "cesfam",
        "migrante",
        "pueblo_originario",
        "vih_positivo",
        "vdrl_resultado",
        "vdrl_tratamiento",
        "hepatitis_b_resultado",
        "profilaxis_completa",
        "creado_por_id",
        "fecha_creacion",
        "estado",
        "fecha_envio",
        "validado_por_id",
        "fecha_validacion",
    ]
    ws_madre.append(headers_madre)

    for m in madres_qs.select_related("creado_por", "validado_por"):
        ws_madre.append([
            m.id,
            m.nombre_completo,
            m.rut,
            remove_tz(m.fecha_nacimiento),
            m.edad,
            m.comuna,
            m.cesfam,
            m.migrante,
            m.pueblo_originario,
            m.vih_positivo,
            m.vdrl_resultado,
            m.vdrl_tratamiento,
            m.hepatitis_b_resultado,
            m.profilaxis_completa,
            m.creado_por_id,
            remove_tz(m.fecha_creacion),
            m.estado,
            remove_tz(m.fecha_envio),
            m.validado_por_id,
            remove_tz(m.fecha_validacion),
        ])

    auto_adjust(ws_madre)

    # =========================
    # HOJA PARTO
    # =========================
    ws_parto = wb.create_sheet("Parto")

    headers_parto = [
        "id",
        "madre_id",
        "fecha_parto",
        "hora_parto",
        "tipo_parto",
        "ive_causal2",
        "acompanante_pabellon",
        "clampeo_tardio",
        "gases_cordon",
        "apego_canguro",
        "apego_tunel",
        "profesional_responsable",
        "observaciones",
        "edad_gestacional_semanas",
        "edad_gestacional_dias",
        "monitor",
        "ttc",
        "induccion",
        "alumno",
        "plan_parto",
        "alumbramiento_rigido",
        "clasificacion_robson",
        "acompanante_parto",
        "motivo_no_acompanado",
        "persona_acompanante",
        "acompanante_corta_cordon",
        "causa_cesarea",
        "uso_sala_saip",
        "recuerdos",
        "retira_placenta",
        "estampado_placenta",
        "manejo_dolor_farmacologico",
    ]
    ws_parto.append(headers_parto)

    partos = Parto.objects.filter(madre__in=madres_qs)
    for p in partos:
        ws_parto.append([
            p.id,
            p.madre_id,
            remove_tz(p.fecha_parto),
            p.hora_parto,
            p.tipo_parto,
            p.ive_causal2,
            p.acompanante_pabellon,
            p.clampeo_tardio,
            p.gases_cordon,
            p.apego_canguro,
            p.apego_tunel,
            p.profesional_responsable,
            p.observaciones,
            p.edad_gestacional_semanas,
            p.edad_gestacional_dias,
            p.monitor,
            p.ttc,
            p.induccion,
            p.alumno,
            p.plan_parto,
            p.alumbramiento_rigido,
            p.clasificacion_robson,
            p.acompanante_parto,
            p.motivo_no_acompanado,
            p.persona_acompanante,
            p.acompanante_corta_cordon,
            p.causa_cesarea,
            p.uso_sala_saip,
            p.recuerdos,
            p.retira_placenta,
            p.estampado_placenta,
            p.manejo_dolor_farmacologico,
        ])

    auto_adjust(ws_parto)

    # =========================
    # HOJA RN
    # =========================
    ws_rn = wb.create_sheet("RecienNacido")

    headers_rn = [
        "id",
        "parto_id",
        "apellido_paterno",
        "sexo",
        "peso",
        "talla",
        "circunferencia_craneana",
        "diagnostico",
        "malformacion_congenita",
        "descripcion_malformacion",
        "apgar_1",
        "apgar_5",
        "reanimacion_basica",
        "reanimacion_avanzada",
        "profilaxis_ocular",
        "vacuna_hepatitis_b",
        "vacuna_bcg",
        "profesional_vacuna_vhb",
        "profilaxis_completa",
        "semanas_gestacion",
        "dias_gestacion",
        "apego_inmediato",
        "destino_final",
        "notas_rn",
    ]
    ws_rn.append(headers_rn)

    rns = RecienNacido.objects.filter(parto__madre__in=madres_qs)
    for rn in rns:
        ws_rn.append([
            rn.id,
            rn.parto_id,
            rn.apellido_paterno,
            rn.sexo,
            rn.peso,
            rn.talla,
            rn.circunferencia_craneana,
            rn.diagnostico,
            rn.malformacion_congenita,
            rn.descripcion_malformacion,
            rn.apgar_1,
            rn.apgar_5,
            rn.reanimacion_basica,
            rn.reanimacion_avanzada,
            rn.profilaxis_ocular,
            rn.vacuna_hepatitis_b,
            rn.vacuna_bcg,
            rn.profesional_vacuna_vhb,
            rn.profilaxis_completa,
            rn.semanas_gestacion,
            rn.dias_gestacion,
            rn.apego_inmediato,
            rn.destino_final,
            rn.notas_rn,
        ])

    auto_adjust(ws_rn)

    # =========================
    # HOJA BCG
    # =========================
    ws_bcg = wb.create_sheet("VacunaBCG")

    headers_bcg = [
        "id",
        "rn_id",
        "numero_registro",
        "nombre_completo_madre",
        "rut_madre",
        "fecha_nacimiento_madre",
        "fecha_nacimiento_hijo",
        "sexo_rn",
        "peso_rn",
        "aplicada",
        "comuna",
        "reaccion_adversa",
        "cama_ubicacion",
    ]
    ws_bcg.append(headers_bcg)

    bcgs = VacunaBCG.objects.filter(rn__parto__madre__in=madres_qs)
    for b in bcgs:
        ws_bcg.append([
            b.id,
            b.rn_id,
            b.numero_registro,
            b.nombre_completo_madre,
            b.rut_madre,
            remove_tz(b.fecha_nacimiento_madre),
            remove_tz(b.fecha_nacimiento_hijo),
            b.sexo_rn,
            b.peso_rn,
            b.aplicada,
            b.comuna,
            b.reaccion_adversa,
            b.cama_ubicacion,
        ])

    auto_adjust(ws_bcg)

    # =========================
    # RESPUESTA FINAL
    # =========================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="formularios_completos.xlsx"'

    wb.save(response)
    return response
